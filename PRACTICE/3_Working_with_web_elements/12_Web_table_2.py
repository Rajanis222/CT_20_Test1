import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By

driver=webdriver.Firefox()
driver.maximize_window()
driver.get("https://www.tutorialspoint.com/selenium/practice/webtables.php")
web_table_rows=driver.find_elements(By.XPATH,"//table[@class='table table-striped mt-3']/tbody/tr")
print(f"No of rows in web table:{len(web_table_rows)}")

# table data
data_list=[]
for row in web_table_rows:
    col=row.find_elements(By.TAG_NAME,"td")
    data=[col[0].text, col[1].text,col[2].text,col[3].text, col[4].text, col[5].text]
    data_list.append(data)
print(data_list)

#table header
web_table_header=driver.find_elements(By.XPATH,"//table[@class='table table-striped mt-3']//th")
print(f"len(web_table_header)={len(web_table_header)}")
header_list=[]
for header in web_table_header:
    header_name=header.text
    header_list.append(header_name)
# print(f"header_list:{header_list}")

#write header in excel file
file=r"D:\CREDENCE_CT20_FILES\SELENIUM\PYTHON_SELENIUM\PRACTICE\3_Working_with_web_elements\webtable.xlsx"
sheet_name="sheet2"

excel_file=openpyxl.load_workbook(file)
sheet=excel_file[sheet_name]

for i in range(len(header_list)):
    sheet.cell(row=1,column=i+1).value=header_list[i]

# write data in excel file
for i in range(len(data_list)):
    for j in range(len(data_list[i])):
        sheet.cell(row=i+2,column=j+1).value=data_list[i][j]

excel_file.save(file)
excel_file.close()
driver.close()
