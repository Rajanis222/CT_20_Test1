import time

import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By

driver=webdriver.Firefox()
driver.maximize_window()

#webtable 1

driver.get("https://apps.credence.in/practice")
webTable_scroll = driver.find_element(By.XPATH, "//legend[normalize-space()='Web Table Example']")      # Click on Prompt Alert" Button
driver.execute_script("arguments[0].scrollIntoView();", webTable_scroll)

#Data
# web_table=driver.find_elements(By.XPATH,"//table[@id='product']/tbody/tr")
# print(len(web_table))
# data=[]
# for row in web_table:
#     # print(web_table.index(row))
#      col=row.find_elements(By.TAG_NAME,"td")
#      print(f"Row no-->{web_table.index(row)+1}, Instructor-->{col[0].text}, Course-->{col[1].text},Price-->{col[2].text}")

# #webtable 2
# driver.get("https://www.tutorialspoint.com/selenium/practice/webtables.php")
# webtable_2=driver.find_elements(By.XPATH,"//table[@class='table table-striped mt-3']/tbody/tr")
# print(len(webtable_2))
#
# for row in webtable_2:
#     #head=row.find_elements(By.TAG_NAME,"th")
#     #print(f"headers-->{head[0].text, head[1].text, head[2].text,head[3].text,head[4].text,head[5].text}")
#     col=row.find_elements(By.TAG_NAME,"td")
#     print(f"Row no--> {webtable_2.index(row)+1},first_name:{col[0].text},last_name:{col[1].text},age:{col[2].text},Email{col[3].text} Salary{col[4].text} Department{col[5].text}")


# #webtable 3
#driver.get("https://www.techlistic.com/2017/02/automate-demo-web-table-with-selenium.html")
# webtable_3=driver.find_elements(By.XPATH,"//table[@id='customers']/tbody/tr[2]")
# rows=len(webtable_3)
# print(rows)
# # data=[]
# for row in webtable_3:
#     cols=row.find_elements(By.TAG_NAME,"td")
#     try:
#         print(f"row no:{webtable_3.index(row)+1}, Company:{cols[0].text},Contact: {cols[1].text},Country:{cols[2].text}")
#     except:
#         print("list index out of range")
#     if len(cols)>=3:
#         Company=cols[0].text
#         Contact=cols[1].text
#         Country=cols[2].text
#         print(f"Row no: {webtable_3.index(row) + 1}, Company: {Company}, Contact: {Contact}, Country: {Country}")
#     else:
#         print(f"Row no: {webtable_3.index(row) + 1} does not have enough columns.")
#     for col in cols:
#         print(col.text)

#webtable 4
# driver.get("https://practice.expandtesting.com/tables")
# webtable_4=driver.find_elements(By.XPATH,"//table[@id='table1']/tbody/tr")
# print(f"Rows in webtable4:{len(webtable_4)}")
# headers_list=[]
# headers_elements=driver.find_elements(By.XPATH,"//table[@id='table2']//tr/th")
# for h in headers_elements:
#     header=h.text
#     print(h)
#     headers_list.append(header)
# print(f"headers_list:{headers_list}")
#
# for row in webtable_4:
#     col=row.find_elements(By.TAG_NAME,"td")
#     print(f"Row no.:{webtable_4.index(row)+1},last_name:{col[0].text},first_name:{col[1].text},email:{col[2].text},Due:{col[3].text}")
#
# file=r"D:\CREDENCE_CT20_FILES\SELENIUM\PYTHON_SELENIUM\PRACTICE\3_Working_with_web_elements\webtable.xlsx"
# sheet_name="sheet4"
# excel_file=openpyxl.load_workbook(file)
# sheet=excel_file[sheet_name]
#
# # Writing the code for header
# for i in range(len(headers_list)):
#     row_no=1
#     column_no=i
# sheet.cell(row=row_no, column=column_no).value=i
#
# sheet.cell(row=1,column=1).value=





driver.quit()