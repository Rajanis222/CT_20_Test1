import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By

driver=webdriver.Firefox()
driver.maximize_window()
driver.get("https://apps.credence.in/practice")

# data
web_table=driver.find_elements(By.XPATH,"(//table[@id='product'])[1]/tbody/tr")
print(f"no of rows:{len(web_table)}")
data_list=[]
for row in web_table:
    col=row.find_elements(By.TAG_NAME,"td")
    #print(col[0].text,col[1].text,col[2].text)
    data_list.append([col[0].text,col[1].text,col[2].text])
#print(f"data list={data_list}")
print(f"len(data_list)={len(data_list)}")

# header
web_table_headers=driver.find_elements(By.XPATH,"(//table[@id='product'])[1]/thead/tr/th")
header_list=[]
for header in web_table_headers:
    header_name=header.text
    header_list.append(header_name)
    # print(f"header name:{header_name}")
print(f"header list:{header_list}")
print(f"len(header_list)={len(header_list)}")

file=r"D:\CREDENCE_CT20_FILES\SELENIUM\PYTHON_SELENIUM\PRACTICE\3_Working_with_web_elements\webtable.xlsx"
sheet_name="sheet1"

excel_file=openpyxl.load_workbook(file)
sheet=excel_file[sheet_name]
#writing header in excel file
for i in range(len(header_list)): #0,1,2
    column_no=i+1
    sheet.cell(row=1,column=column_no).value=header_list[i]
print()
#writing data in excel file
for i in range(len(data_list)): #0,1,2,3,....9
    for j in range(len(data_list[i])): #0,1,2
        sheet.cell(row=i+2, column=j+1).value=data_list[i][j]
excel_file.save(file)
excel_file.close()
driver.close()






