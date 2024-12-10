import time

import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By

driver=webdriver.Chrome()
driver.maximize_window()
driver.get("https://www.worldometers.info/coronavirus/#countries")
print(f"Title of this page is:{driver.title}")
web_table_row=driver.find_elements(By.XPATH,"//table[@id='main_table_countries_today']/tbody/tr")
print(f"No.of rows in web table:{len(web_table_row)}")

#data
data_list=[]
for row in web_table_row:
    col=row.find_elements(By.TAG_NAME,"td")
    time.sleep(0.3)
    data_list.append([col[0].text,col[1].text,col[2].text,col[3].text,col[4].text,col[5].text,col[6].text,col[7].text])
    # print(f"#: {col[0].text},country,other:{col[1].text}, Total cases:{col[2].text}, Total deaths:{col[3].text},Total recovered:{col[4].text},tot case/1M pop:{col[5].text}, Deaths:{col[6].text},population:{col[7].text}")
#print(data_list)
print(f"len(data_list)={len(data_list)}")

# Header
web_table_header=driver.find_elements(By.XPATH,"//table[@id='main_table_countries_today']//th")
header_list=[]
for header in web_table_header:
    header_name=header.text
    header_list.append(header_name)
# print(header_list)
print(f"len(header_list)={len(header_list)}")

# write header in excel file
file=r"D:\CREDENCE_CT20_FILES\SELENIUM\PYTHON_SELENIUM\PRACTICE\3_Working_with_web_elements\webtable.xlsx"
sheet_name="sheet3"

excel_file=openpyxl.load_workbook(file)
sheet=excel_file[sheet_name]
for i in range(len(header_list)):  #0,1,2,3...8
    sheet.cell(row=1,column=i+1).value=header_list[i]

#write data in excel file
for i in range(len(data_list)): #0,1,2.....246
    for j in range(len(data_list[i])):
        sheet.cell(row=i+2, column=j+1).value=data_list[i][j]
excel_file.save(file)
excel_file.close()
driver.close()

