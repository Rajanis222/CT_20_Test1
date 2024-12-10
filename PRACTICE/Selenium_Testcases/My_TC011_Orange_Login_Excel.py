import openpyxl
from selenium import  webdriver
from selenium.webdriver.common.by import By

username_list = ["Admin", "admin@yourstore.com" , "admin1@yourstore.com","Admin", "admin2", "admin3", "admin4"]
password_list = ["admin123", "admin", "admin1", "admin123", "admin2", "admin3", "admin4"]

# create excel file
file=r"D:\CREDENCE_CT20_FILES\SELENIUM\PYTHON_SELENIUM\PRACTICE\Selenium_Testcases\Test_case.xlsx"
sheet_name="TC_011"
excel_file=openpyxl.load_workbook(file)
sheet=excel_file[sheet_name]
# write header
sheet.cell(row=1,column=1).value="username"
sheet.cell(row=1,column=2).value="password"

#write data in excel file
for i in range(len(username_list)):
    sheet.cell(row=(i+2),column=1).value=(username_list[i])
    sheet.cell(row=(i+2), column=(2)).value=(password_list[i])
excel_file.save(file)
excel_file.close()

# # access the data from excel file
row_count=sheet.max_row
driver=webdriver.Firefox()
driver.maximize_window()
driver.get("https://opensource-demo.orangehrmlive.com/web/index.php/auth/login")
driver.implicitly_wait(10)
# username
for i in range(row_count):
    username_value = sheet.cell(row=i + 2, column=1).value
    username = driver.find_element(By.XPATH, "//input[@placeholder='Username']")
    username.clear()
    username.send_keys("username_value")
# password
    password_value = sheet.cell(row=i + 2, column=2).value
    password=driver.find_element(By.XPATH,"//input[@placeholder='Password']")
    password.clear()
    password.send_keys("password_value")

    login_button=driver.find_element(By.XPATH,"//button[normalize-space()='Login']")
    login_button.click()

    try:
        menu_button=driver.find_element(By.XPATH,"//p[@class='oxd-userdropdown-name']")
        print(f"user {username_value} logged in successfully")
        menu_button.click()
        logout_button=driver.find_element(By.XPATH,"//a[normalize-space()='Logout']")
        logout_button.click()
    except:
        print(f"User {username_value} failed to login")
driver.refresh()
